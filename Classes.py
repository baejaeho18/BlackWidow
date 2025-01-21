from selenium import webdriver
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import (StaleElementReferenceException,
                                       TimeoutException,
                                       UnexpectedAlertPresentException,
                                       NoSuchFrameException,
                                       NoAlertPresentException,
                                       WebDriverException,
                                       InvalidElementStateException
                                       )

from urllib.parse import urlparse, urljoin
import json
import pprint
import datetime
import tldextract
import math
import os
import traceback
import random
import re
import time
import itertools
import string

from bs4 import BeautifulSoup
import hashlib
import adblockparser
import requests
import csv
import matplotlib.pyplot as plt
import networkx as nx

from Functions import *
from extractors.Events import extract_events
from extractors.Forms import extract_forms, parse_form
from extractors.Urls import extract_urls
from extractors.Iframes import extract_iframes
from extractors.Ui_forms import extract_ui_forms

from Elements import Request, Form, Ui_form, Graph, Event, Iframe

import logging
log_file = os.path.join(os.getcwd(), 'logs', 'crawl-'+str(time.time())+'.log')
logging.basicConfig(filename=log_file, format='%(asctime)s\t%(name)s\t%(levelname)s\t[%(filename)s:%(lineno)d]\t%(message)s', datefmt='%Y-%m-%d:%H:%M:%S', level=logging.DEBUG)
# Turn off all other loggers that we don't care about.
for v in logging.Logger.manager.loggerDict.values():
    v.disabled = True


class Crawler:
    def __init__(self, driver, url, timeout=900, output="script_logs.xlsx"):
        self.driver = driver
        # Start url
        self.url = url
        self.graph = Graph()
        self.session_id = str(time.time()) + "-" + str(random.randint(1,10000000))

        logging.info("Init crawl on " + url)

        self.max_depth = 3
        self.max_events = 2
        self.max_forms = 1
        self.max_gets = 1

        self.hash_set = set()
        self.filters = {}

        self.timeout = timeout
        self.start_time = None
        
        self.output_file = output

    def start(self, debug_mode=True):
        self.root_req = Request("ROOTREQ", "get")
        req = Request(self.url, "get")
        self.graph.add(self.root_req)
        self.graph.add(req)
        self.graph.connect(self.root_req, req, CrawlEdge("get", None, None) )
        self.debug_mode = debug_mode
        self.start_time = time.time()

        filter_lists_path = os.path.join("filterlists")
        self.filters = self.get_filter_lists(filter_lists_path)
        
        self.graph.data['urls'] = {}
        self.graph.data['form_urls'] = {}

        random.seed( 6 ) # chosen by fair dice roll

        processed_edges_count = 0
        still_work = True
        while still_work:
            if self.timeout and (time.time() - self.start_time > self.timeout):
                break
            print("-"*50)
            new_edges = len([edge for edge in self.graph.edges if edge.visited == False])
            print("Edges left: %s | Processed: %s" % (new_edges, processed_edges_count))  # 처리된 Edge 출력
            try:
                n_gets = 0
                n_forms = 0
                n_events = 0
                for edge in self.graph.edges:
                    if edge.visited == False:
                        if edge.value.method == "get":
                            n_gets += 1
                        elif edge.value.method == "form":
                            n_forms += 1
                        elif edge.value.method == "event":
                            n_events += 1
                print()
                print("----------------------")
                print("GETS    | FROMS  | EVENTS ")
                print(str(n_gets).ljust(7), "|", str(n_forms).ljust(6), "|", n_events)
                print("----------------------")

                # for edge in self.graph.edges:
                #     if edge.visited == False and edge.value.method=="get":
                #         print(edge)
                # input()

                try:
                    still_work = self.fine_crawl()
                    processed_edges_count += 1
                except Exception as e:
                    still_work = n_gets + n_forms + n_events
                    print(e)
                    print(traceback.format_exc())
                    logging.error(e)

                    logging.error("Top level error while crawling")
                #input("Enter to continue")

            except KeyboardInterrupt:
                print("CTRL-C, abort mission")
                #print(self.graph.toMathematica())
                break

        print("Done crawling")

    # Handle priority
    def next_unvisited_edge(self, driver, graph):
        for edge in graph.edges:
            if not edge.visited:
                if check_edge(driver, graph, edge) and follow_edge(driver, graph, edge):
                    return edge
                else:
                    logging.warning("Edge check or follow failed: " + str(edge))
                    edge.visited = True

        return None

    def load_page(self, driver, graph):
        request = None
        edge = self.next_unvisited_edge(driver, graph)
        if not edge:
            return None

        # Update last visited edge
        graph.data['prev_edge'] = edge

        request = edge.n2.value

        logging.info("Current url: " + driver.current_url)
        logging.info("Crawl (edge): " +  str(edge) )
        print("Crawl (edge): " +  str(edge) )

        return (edge,request)

    # Actually not recursive (TODO change name)
    def fine_crawl(self):
        driver = self.driver
        graph = self.graph

        todo = self.load_page(driver, graph)
        if not todo:
            print("Done crawling")
            print(graph)

            f = open("graph_mathematica.txt", "w")
            f.write( self.graph.toMathematica() )

            return False

        (edge, request) = todo
        graph.visit_node(request)
        graph.visit_edge(edge)

        # (almost) Never GET twice (optimization)
        if edge.value.method == "get":
            for e in graph.edges:
                if (edge.n2 == e.n2) and (edge != e) and (e.value.method == "get"):
                    #print("Fake visit", e)
                    graph.visit_edge(e)

        # Wait if needed
        try:
            wait_json = driver.execute_script("return JSON.stringify(need_to_wait)")
            wait = json.loads(wait_json)
            if wait:
                time.sleep(1)
        except UnexpectedAlertPresentException:
            logging.warning("Alert detected")
            alert = driver.switch_to_alert()
            alert.dismiss()

            # Check if double check is needed...
            try:
                wait_json = driver.execute_script("return JSON.stringify(need_to_wait)")
                wait = json.loads(wait_json)
                if wait:
                    time.sleep(1)
            except:
                logging.warning("Inner wait error for need_to_wait")
        except:
            logging.warning("No need_to_wait")

        # Timeouts
        try:
            resps = driver.execute_script("return JSON.stringify(timeouts)")
            todo = json.loads(resps)
            for t in todo:
                try:
                    if t['function_name']:
                        driver.execute_script(t['function_name'] + "()")
                except:
                    logging.warning("Could not execute javascript function in timeout " + str(t))
        except:
            logging.warning("No timeouts from stringify")

        self.extract_JS(edge)

        print(f"Depth: {edge.depth}/{self.max_depth}")
        print(f"Events:{edge.cnt_gets}/{self.max_events}")
        print(f"Forms: {edge.cnt_forms}/{self.max_forms}")
        print(f"Gets : {edge.cnt_gets}/{self.max_gets}")

        if edge.depth > self.max_depth:
            return True
        
        # Check if we need to wait for asynch
        try:
            wait_json = driver.execute_script("return JSON.stringify(need_to_wait)")
        except UnexpectedAlertPresentException:
            logging.warning("Alert detected")
            alert = driver.switch_to_alert()
            alert.dismiss()
        wait_json = driver.execute_script("return JSON.stringify(need_to_wait)")
        wait = json.loads(wait_json)
        if wait:
            time.sleep(1)

        # Add findings to the graph
        current_cookies = driver.get_cookies()

        if edge.cnt_gets <= self.max_gets:
            reqs = extract_urls(driver)
            logging.info("Adding requests from URLs")
            for req in reqs:
                logging.info("from URLs %s " % str(req))
                new_edge = graph.create_edge(request, req, CrawlEdge(req.method, None, current_cookies), edge )
                if allow_edge(graph, new_edge):
                    graph.add(req)
                    graph.connect(request, req, CrawlEdge(req.method, None, current_cookies), edge )
                else:
                    logging.info("Not allowed to add edge: %s" % new_edge)

        if edge.cnt_forms <= self.max_forms:
            forms = extract_forms(driver)
            forms = set_form_values(forms)
            ui_forms = extract_ui_forms(driver)
            logging.info("Adding requests from froms")
            for form in forms:
                req = Request(form.action, form.method)
                logging.info("from forms %s " % str(req))
                new_edge = graph.create_edge( request, req, CrawlEdge("form", form, current_cookies), edge )
                if allow_edge(graph, new_edge):
                    graph.add(req)
                    graph.connect(request, req, CrawlEdge("form", form, current_cookies), edge )
                else:
                    logging.info("Not allowed to add edge: %s" % new_edge)
            logging.info("Adding requests from ui_forms")
            for ui_form in ui_forms:
                req = Request(driver.current_url, "ui_form")
                logging.info("from ui_forms %s " % str(req))

                new_edge = graph.create_edge( request, req, CrawlEdge("ui_form", ui_form, current_cookies), edge )
                if allow_edge(graph, new_edge):
                    graph.add(req)
                    graph.connect(request, req, CrawlEdge("ui_form", ui_form, current_cookies), edge )
                else:
                    logging.info("Not allowed to add edge: %s" % new_edge)

        if edge.cnt_events <= self.max_events:
            events = extract_events(driver)
            logging.info("Adding requests from events")
            for event in events:
                req = Request(request.url, "event")
                logging.info("from events %s " % str(req))

                new_edge = graph.create_edge( request, req, CrawlEdge("event", event, current_cookies), edge )
                if allow_edge(graph, new_edge):
                    graph.add(req)
                    graph.connect(request, req, CrawlEdge("event", event, current_cookies), edge )
                else:
                    logging.info("Not allowed to add edge: %s" % new_edge)

        iframes = extract_iframes(driver)
        logging.info("Adding requests from iframes")
        for iframe in iframes:
            req = Request(iframe.src, "iframe")
            logging.info("from iframes %s " % str(req))

            new_edge = graph.create_edge( request, req, CrawlEdge("iframe", iframe, current_cookies), edge )
            if allow_edge(graph, new_edge):
                graph.add(req)
                graph.connect(request, req, CrawlEdge("iframe", iframe, current_cookies), edge )
            else:
                logging.info("Not allowed to add edge: %s" % new_edge)

        # Try to clean up alerts
        try:
            alert = driver.switch_to_alert()
            alert.dismiss()
        except NoAlertPresentException:
            pass

        return True

    # 3) Extract JS
    def hash_content(self, content):
        return hashlib.sha256(content.encode('utf-8')).hexdigest()

    def get_filter_lists(self, filter_lists_path):
        if not os.path.exists(filter_lists_path):
            os.makedirs(filter_lists_path)
        
        filter_lists = {
            'EasyList': 'https://easylist.to/easylist/easylist.txt',
            'EasyPrivacy': 'https://easylist.to/easylist/easyprivacy.txt',
            'Fanboy\'s Annoyance List': 'https://easylist.to/easylist/fanboy-annoyance.txt',
            'Peter Lowe\'s List': 'https://pgl.yoyo.org/adservers/serverlist.php?hostformat=adblockplus&mimetype=plaintext',
        }
        for name, url in filter_lists.items():
            filter_list_path = os.path.join(filter_lists_path, f'{name}.txt')
            
            # 파일이 존재하면 다운로드하지 않고 읽기만 함
            if not os.path.exists(filter_list_path):
                print(f"Downloading {name} filter list from {url}")
                content = requests.get(url).content
                with open(filter_list_path, 'wb') as f:
                    f.write(content)

            # 필터 목록 파일 읽기
            with open(filter_list_path, 'r') as f:
                raw_rules = f.read().splitlines()
            
            # AdblockRules 객체로 저장
            filter_lists[name] = adblockparser.AdblockRules(raw_rules)
        
        return filter_lists

    def is_tracker(self, script_url):
        checked_filters = []
        for name, filter_obj in self.filters.items():
            if filter_obj.should_block(script_url):
                checked_filters.append(name)
        return len(checked_filters) > 0, checked_filters

    def save_to_excel(self, excel_file_path, root_url, trigger_sequence, script_url, content_hash, label, checked_filter):
        from openpyxl import Workbook, load_workbook
        
        if not os.path.exists(excel_file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "JavaScript Logs"
            ws.append(["Root URL", "Trigger Sequence", "Depth", "Script URL", "Content Hash", "Label", "Checked Filter"])
            wb.save(excel_file_path)
        
        wb = load_workbook(excel_file_path)
        ws = wb.active
        ws.append([
            root_url,
            "-> ".join([f"{edge.n1.value} -( {edge.value} )" for edge in trigger_sequence] + [str(trigger_sequence[-1].n2.value)]),
            len(trigger_sequence),
            script_url,
            content_hash,
            label,
            ", ".join(checked_filter) if isinstance(checked_filter, list) else checked_filter
        ])
        wb.save(excel_file_path)

    def save_file(self, file_path, content):
        with open(file_path, 'a', encoding='utf-8') as f:
            f.write(content)

    def extract_JS(self, edge):
        OUTPUT_DIR = "./js_output"
        driver = self.driver
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        load_js = 0

        # 시퀀스 생성: parent를 따라 올라가며 시퀀스 구성
        trigger_sequence = []
        current_edge = edge
        while current_edge is not None:
            trigger_sequence.append(current_edge)
            current_edge = current_edge.parent
        trigger_sequence = trigger_sequence[::-1]  # 시퀀스를 올바른 순서로 정렬

        # Inline JavaScript 수집
        # inline_js = [tag.text for tag in soup.find_all('script') if tag.text.strip()]
        # for script in inline_js:
        #     load_js += self.save_script(script, "Inline", trigger_sequence, OUTPUT_DIR)

        # External JavaScript 수집 (DOM 기반)
        external_js_urls = [tag['src'] for tag in soup.find_all('script', src=True)]
        for script_url in external_js_urls:
            full_script_url = urljoin(driver.current_url, script_url)
            try:
                response = requests.get(full_script_url)
                if response.status_code == 200:
                    load_js += self.save_script(response.text, full_script_url, trigger_sequence, OUTPUT_DIR)
                else:
                    print(f"Failed to fetch script: {full_script_url}, status code: {response.status_code}")
            except requests.RequestException as e:
                print(f"Error fetching script: {full_script_url}, error: {e}")

        # AdCPG 방식 추가: 네트워크 트래픽 기반 수집
        script_entries = driver.execute_script("""
            return performance.getEntriesByType('resource')
                .filter(entry => entry.initiatorType === 'script')
                .map(entry => entry.name);      
        """)
        for script_url in script_entries:
            try:
                response = requests.get(script_url)
                if response.status_code == 200:
                    load_js += self.save_script(response.text, script_url, trigger_sequence, OUTPUT_DIR)
                else:
                    print(f"Failed to fetch script: {script_url}, status code: {response.status_code}")
            except requests.RequestException as e:
                print(f"Error fetching script: {script_url}, error: {e}")

        # 로드된 JS가 없을 경우 기록
        if load_js == 0:
            self.save_to_excel(
                os.path.join(OUTPUT_DIR, "script_logs.xlsx"),
                self.url,
                trigger_sequence,
                "None",  # JS가 로드되지 않았음을 표시
                "N/A",
                "-",
                []
            )
        trigger_sequence[-1].n2.update_js_info(load_js)

    # 공통 스크립트 저장 메서드
    def save_script(self, content, url_or_label, trigger_sequence, output_dir):
        
        if isinstance(url_or_label, str) and url_or_label == "Inline":
            hash = self.hash_content(content)
        else:
            hash = self.hash_content(url_or_label)

        if hash in self.hash_set:
            return 0
        self.hash_set.add(hash)
        
        script_dir = os.path.join(output_dir, hash)
        # Labeling
        if url_or_label == "Inline":
            label = "Inline"
            checked_filters = []
        else:
            is_tracking, checked_filters = self.is_tracker(url_or_label)
            label = "1" if is_tracking else "0"
            
        if not os.path.exists(script_dir):
            os.makedirs(script_dir)
            self.save_file(os.path.join(script_dir, "code"), content)
            self.save_file(os.path.join(script_dir, "label"), label)
            return 1
        
        self.save_to_excel(
                os.path.join(output_dir, self.output_file),
                self.url,
                trigger_sequence,
                url_or_label if url_or_label != "Inline" else f"url_inline_{script_dir}",
                hash,
                label,
                checked_filters
        )
        return 0


    # 4) Graph Representation
    def graph_visualizer(self):
        # NetworkX 그래프 객체 생성
        G = nx.DiGraph()  # 방향성이 있는 그래프 (일반적으로 웹 크롤링에서는 방향성이 중요)

        # 그래프에 노드 추가
        for node in self.graph.nodes:
            G.add_node(node.value)  # Node의 value를 사용하여 노드를 추가

        # 그래프에 엣지 추가
        for edge in self.graph.edges:
            G.add_edge(edge.n1.value, edge.n2.value, method=edge.value)

        # 그래프의 레이아웃을 설정
        pos = nx.spring_layout(G)  # spring_layout은 노드 간의 거리를 자연스럽게 배치해줍니다.

        # 노드와 엣지를 시각화
        plt.figure(figsize=(12, 12))  # 그래프의 크기를 설정
        nx.draw(G, pos, with_labels=True, node_size=3000, node_color='lightblue', font_size=10, font_weight='bold', edge_color='gray')

        # 엣지의 레이블 (요청 메서드 표시)
        edge_labels = nx.get_edge_attributes(G, 'method')
        nx.draw_networkx_edge_labels(G, pos, edge_labels=edge_labels)

        # 그래프 표시
        plt.title("Graph Visualization")
        plt.show()


# Edge with specific crawling info, cookies, type of request etc.
class CrawlEdge:
    def __init__(self, method, method_data, cookies):
        self.method = method
        self.method_data = method_data
        self.cookies = cookies

    def __repr__(self):
        return str(self.method) + " " + str(self.method_data)
    # Cookies are not considered for equality.
    def __eq__(self, other):
        return (self.method == other.method and self.method_data == other.method_data)

    def __hash__(self):
        return hash( hash(self.method) + hash(self.method_data) )